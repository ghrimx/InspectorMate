import logging
import json
from datetime import datetime, timezone
from qtpy import QtCore, Signal, Slot, QtSql, QSqlRelationalTableModel
from pathlib import Path

from functools import partial
from database.database import AppDatabase

from common import DatabaseField, Signage, Connector, SignageType, SignageStatus, UpdateItem, OETag

from base_models import TreeItem, TreeModel, ProxyModel
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers, PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from html2text import html2text
from utilities.utils import mergeExcelFiles, find_match, extract_hash_lines

from onenote.model import getTags

from utilities.config import config as mconf

logger = logging.getLogger(__name__)


class WorkerSignals(QtCore.QObject):
    finished = Signal(object, str)
    error = Signal(Exception)
    result = Signal(object)
    progress = Signal(int)


class LoadWorker(QtCore.QRunnable):
    def __init__(self, func: callable, *args, cache=None, **kwargs):
        super().__init__()
        self.func = func
        self.args = args
        self.kwargs = kwargs
        self.cache = cache
        self.signals = WorkerSignals()
    
    def run(self):
        try:
            for item in self.func(*self.args, **self.kwargs):
                self.signals.result.emit(item)
        except Exception as e:
            self.signals.error.emit(e)
            msg = "⚠️ Error while loading data from connector!"
        else:
            msg = "✔️ Data imported from connector!"
        finally:
            self.signals.finished.emit(self.cache, msg)


class ExportWorker(QtCore.QRunnable):
    def __init__(self, sql_model, types, statuses, destination, include_publicnote):
        super().__init__()
        self.sql_model: "SignageSqlModel" = sql_model
        self.types: list = types
        self.statuses: list = statuses
        self.destination: str = destination
        self.include_publicnote: bool = include_publicnote
        self.signals = WorkerSignals()

    def func(self):
        """Export Signage to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "main"

        if self.include_publicnote:
            headers = ["Refkey", "Title", "Status", "Owner", "Type", "Evidence", "Note"]
            xrange = "A1:G"
            model_fields = [SignageSqlModel.Fields.Refkey.index,
                            SignageSqlModel.Fields.Title.index,
                            SignageSqlModel.Fields.Status.index,
                            SignageSqlModel.Fields.Owner.index,
                            SignageSqlModel.Fields.Type.index,
                            SignageSqlModel.Fields.DocCount.index,
                            SignageSqlModel.Fields.PublicNote.index]
        else:
            headers = ["Refkey", "Title", "Status", "Owner", "Type", "Evidence"]
            xrange = "A1:F"
            model_fields = [SignageSqlModel.Fields.Refkey.index,
                            SignageSqlModel.Fields.Title.index,
                            SignageSqlModel.Fields.Status.index,
                            SignageSqlModel.Fields.Owner.index,
                            SignageSqlModel.Fields.Type.index,
                            SignageSqlModel.Fields.DocCount.index]

        for column in range(len(headers)):
            ws.cell(row=1, column=column + 1, value=headers[column])

        # Selected Status List
        # Select All if user did not made any choice
        if len(self.statuses) == 0:
            value: SignageStatus
            for value in AppDatabase.cache_signage_status.values():
                self.statuses.append(value.name)

        record_count = 1
        for row in range(self.sql_model.rowCount()):
            record = self.sql_model.record(row)
            if record.value(SignageSqlModel.Fields.Type.index) in self.types:
                if record.value(SignageSqlModel.Fields.Status.index) in self.statuses:
                    record_count += 1
                    values = list(map(record.value, model_fields))

                    # Signage Type
                    values[4] = record.value(SignageSqlModel.Fields.Type.index)

                    # Signage Status
                    values[2] = record.value(SignageSqlModel.Fields.Status.index)
    
                    if self.include_publicnote:
                        if values[-1] is not None:
                            values[-1] = html2text(values[-1]).strip()

                    ws.append(values)

        if record_count == 1:
            record_count = 2

        table = Table(displayName="Table1", ref=f"{xrange}{record_count}")
        style = TableStyleInfo(name="TableStyleMedium2",
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=False)
        table.tableStyleInfo = style

        ws.column_dimensions["B"].width = 150.0
        ws.column_dimensions["A"].number_format = numbers.FORMAT_TEXT

        # Conditional formatting
        red_text = Font(color="9C0006")
        red_fill = PatternFill(bgColor="FFC7CE")
        ws.conditional_formatting.add(f"F2:F{record_count}", CellIsRule(operator='equal', formula=['0'], stopIfTrue=False, font=red_text, fill=red_fill))

        for cell in ws['F':'F']:
            cell.alignment = Alignment(horizontal="center")

        ws.add_table(table)
        wb.save(self.destination)

    def run(self):
        try:
            self.func()
        except Exception as e:
            self.signals.error.emit(e)
            msg = "⚠️ Error while exporting data to Excel"
        else:
            msg = "✔️ Data exported to Excel!"
        finally:
            self.signals.finished.emit(msg, None)


class ExcelLoader(QtCore.QRunnable):
    class Signals(QtCore.QObject):
        batchReady = Signal(list)   # List[UpdateItem]
        signageBatch = Signal(list) # List[Signage]
        finished = Signal(str)
        error = Signal(Exception)

    def __init__(self, sql_model, selected_files, update_title, batch_size=100):
        super().__init__()
        self.sql_model: "SignageSqlModel" = sql_model
        self.selected_files = selected_files
        self.update_title = update_title
        self.batch_size = batch_size
        self.signals = self.Signals()

    def func(self):
        from pandas import DataFrame
        df: DataFrame = mergeExcelFiles(self.selected_files, drop_duplicate="first", outfile="")
        
        try:
            df["Type"]
            df["Title"]
            df["Owner"]
            df['Refkey']
        except KeyError as e:
            raise KeyError(f"Missing required column header: {e}") 
        except Exception as e:
            raise Exception(f"Error while validating file format: {e}")
        
        df.sort_values(by=['Refkey'], inplace=True)
        df = df.fillna("", axis=0)

        updates: list[UpdateItem] = []
        insertions: list[Signage] = []

        self.sql_model.sort(self.sql_model.Fields.Refkey.index, QtCore.Qt.SortOrder.AscendingOrder)

        df["Refkey"] = df["Refkey"].str.lower()
        df["Type"] = df["Type"].str.lower()
        for _, df_row in df.iterrows():
            df_refkey: str = df_row["Refkey"]
            df_type: str = df_row["Type"]
            df_title: str = df_row["Title"]
            df_owner: str = df_row["Owner"]

            found = False
            cache_signage_type = AppDatabase.cache_signage_type.get(df_type.capitalize())
            if cache_signage_type:
                for row in range(self.sql_model.rowCount()):
                    refkey_index = self.sql_model.index(row, self.sql_model.Fields.Refkey.index)
                    m_refkey: str = self.sql_model.data(refkey_index, QtCore.Qt.ItemDataRole.EditRole)
                    type_index = self.sql_model.index(row, self.sql_model.Fields.Type.index)
                    m_signage_type: str = self.sql_model.data(type_index, QtCore.Qt.ItemDataRole.EditRole)
                    title_index = self.sql_model.index(row, self.sql_model.Fields.Title.index)
                    m_title: str = self.sql_model.data(type_index, QtCore.Qt.ItemDataRole.EditRole)
                    
                    if m_refkey.lower() == df_refkey and m_signage_type.lower() == df_type:
                        found = True
                        if self.update_title and (df_title.strip().lower() != m_title.strip().lower()):
                            signage_id = self.sql_model.index(row, self.sql_model.Fields.ID.index).data()
                            updates.append(UpdateItem(signage_id, df_title))
                        break

            if not found and df_refkey and cache_signage_type:
                signage = Signage(
                    refkey=df_refkey,
                    title=df_title,
                    owner=df_owner,
                    type=cache_signage_type.uid if cache_signage_type else None,
                    source=f'{{"application":"InspectorMate", "module":"loadFromExcel"}}',
                    workspace_id=AppDatabase.activeWorkspace().id,
                )
                insertions.append(signage)

            # Emit in batches to avoid memory buildup
            if len(updates) >= self.batch_size:
                self.signals.batchReady.emit(updates.copy())
                updates.clear()
            if len(insertions) >= self.batch_size:
                self.signals.signageBatch.emit(insertions.copy())
                insertions.clear()

        # Emit remaining
        if updates:
            self.signals.batchReady.emit(updates)
        if insertions:
            self.signals.signageBatch.emit(insertions)
    
    @Slot()
    def run(self):
        try:
            self.func()
        except Exception as e:
            self.signals.error.emit(e)
            msg = "⚠️ Error while loading data from Excel file(s)"
        else:
            msg = "✔️ Excel file(s) loaded!"
        finally:
            self.signals.finished.emit(msg)


class DataService:
    @staticmethod
    def loadFromDocx(connectors: dict,
                     regex: str,
                     cache: dict,
                     on_ready: callable,
                     on_finished: callable = None):

        def func(connectors: dict, regex: str, cache: dict):

            connector: Connector
            for connector in connectors.values():
                fpath = Path(connector.value)

                if not fpath.is_file():
                    raise FileNotFoundError

                last_modified = fpath.stat().st_mtime_ns
                if connector.last_modified != last_modified:
                    connector.last_modified = last_modified
                    result = extract_hash_lines(connector.value)
                    
                    for lid, line in result.items():
                        if lid in cache.get("Docx"):
                            continue
                        text = line[1:]
                        signage = Signage()
                        signage.title = html2text(text).strip()
                        signage.refkey = find_match(text, regex)
                        signage.type = 3 if line[0] == '!' else 0
                        signage.workspace_id = AppDatabase.activeWorkspace().id
                        src = (f'{{"application":"Docx", "module":"loadFromDocx",' 
                               f'"file":"{connector.value}", "object_id":"{lid}"}}')
                        signage.source = src
                        cache.get("Docx").add(lid)
                        yield signage

        pool = QtCore.QThreadPool().globalInstance()

        worker = LoadWorker(partial(func, connectors, regex, cache), cache=cache)
        worker.signals.result.connect(on_ready)
        worker.signals.finished.connect(on_finished)
        worker.signals.error.connect(lambda e: logger.error(e))
        pool.start(worker)

    @staticmethod
    def loadFromOneNote(connectors: dict,
                        regex: str,
                        cache: dict,
                        on_ready: callable,
                        on_finished: callable):
        
        pool = QtCore.QThreadPool().globalInstance()

        def func(connectors: dict, regex: str, cache: dict):

            connector: Connector
            for connector in connectors.values():
                section_id = connector.value
                section_name = connector.name
                
                ps_script = mconf.app_data_path.joinpath("onenotescrapper.ps1").as_posix()
                tags = getTags(ps_script, section_id)                

                tag: OETag
                for tag in tags:
                    if tag.ID in cache.get("OneNote"):
                        continue
                    signage = Signage()
                    signage.title = tag.Text
                    signage.refkey = find_match(tag.Text, regex)
                    signage_type: SignageType = AppDatabase.cache_signage_type.get(tag.TypeName.capitalize().strip())

                    # Ignore unknown signage
                    if signage_type is None:
                        logger.debug(f"Unknown tag's type: {tag}")
                        continue

                    signage.type = signage_type.uid
                    signage.workspace_id = AppDatabase.activeWorkspace().id
                    src = (f'{{"application":"OneNote", "module":"loadFromOnenote",'
                           f'"section":"{section_name}", "page":"{tag.PageName}",'
                           f'"object_id":"{tag.ID}"}}')
                    signage.source = src
                    signage.creation_datetime = (datetime.fromisoformat(tag.CreationTime[:-1])
                                                 .astimezone(timezone.utc).strftime('%Y-%m-%d'))
                    signage.modification_datetime = tag.LastModifiedTime
                    cache.get("OneNote").add(tag.ID)
                    yield signage
  
        worker = LoadWorker(partial(func, connectors, regex, cache), cache=cache)
        worker.signals.result.connect(on_ready)
        worker.signals.finished.connect(on_finished)
        worker.signals.error.connect(lambda e: logger.error(e))
        pool.start(worker)

    @staticmethod
    def export2Excel(sql_model: "SignageSqlModel",
                     types: list,
                     statuses: list,
                     destination: str,
                     include_publicnote: bool,
                     on_finished: callable):
        
        pool = QtCore.QThreadPool().globalInstance()
        worker = ExportWorker(sql_model, types, statuses, destination, include_publicnote)
        worker.signals.finished.connect(on_finished)
        worker.signals.error.connect(lambda e: logger.error(e))
        pool.start(worker)

    @staticmethod
    def loadFromExcel(model: "SignageModel",
                      selected_files,
                      update_title,
                      on_signage_ready,
                      stopSpinner):

        sql_model = model.rootModel()
        loader = ExcelLoader(sql_model, selected_files, update_title)

        def applyBatch(updates: list[UpdateItem]):
            model.layoutAboutToBeChanged.emit()
            for upd in updates:
                index: QtCore.QModelIndex = model.findIndexById(upd.signage_id, SignageSqlModel.Fields.ID.index)
                if index.isValid():
                    title_index = index.sibling(index.row(), SignageSqlModel.Fields.Title.index)
                    model.setData(title_index, upd.title, QtCore.Qt.ItemDataRole.EditRole)
            model.layoutChanged.emit()

        def applyInsertions(signages: list[Signage]):
            for signage in signages:
                on_signage_ready(signage)

        loader.signals.batchReady.connect(applyBatch)
        loader.signals.signageBatch.connect(applyInsertions)
        loader.signals.finished.connect(stopSpinner)
        loader.signals.error.connect(lambda e: logger.error(e))

        QtCore.QThreadPool.globalInstance().start(loader)


class SignageSqlModel(QSqlRelationalTableModel):

    class Fields:
        Refkey: DatabaseField
        Title: DatabaseField
        Owner: DatabaseField
        Type: DatabaseField
        Status: DatabaseField
        Source: DatabaseField
        Note: DatabaseField
        PublicNote: DatabaseField
        CreationDatetime: DatabaseField
        ModificationDatetime: DatabaseField
        ID: DatabaseField
        ParentID: DatabaseField
        Workspace: DatabaseField
        DocCount: DatabaseField # virtual column
        Progress: DatabaseField # virtual column

        @classmethod
        def fields(self) -> list["DatabaseField"]:
            """Return all defined DatabaseField instances."""
            return [
                value for name, value in self.__dict__.items()
                if isinstance(value, DatabaseField)
            ]
        
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setTable("signage")
        self.initFields()  

        self.setRelation(self.fieldIndex("type"),
                         QtSql.QSqlRelation("signage_type", "uid", "name"))
        self.setRelation(self.fieldIndex("status"),
                         QtSql.QSqlRelation("signage_status", "uid", "name"))
        
        self._renameHeaders()

        self.setEditStrategy(QtSql.QSqlTableModel.EditStrategy.OnFieldChange)
        self.setFilter(f"workspace_id={AppDatabase.activeWorkspace().id}")
        self.select()

    def columnCount(self, parent = QtCore.QModelIndex()):
        """Add virtual column"""
        return super().columnCount(parent) + 2
    
    def initFields(self):
        self.Fields.ID = DatabaseField('signage_id', self.fieldIndex('signage_id'), False)
        self.Fields.ParentID = DatabaseField('parentID', self.fieldIndex('parentID'), False)
        self.Fields.Status = DatabaseField('Status', self.fieldIndex('status'), True)
        self.Fields.Refkey = DatabaseField('Refkey', self.fieldIndex('refkey'), True)
        self.Fields.Title = DatabaseField('Title', self.fieldIndex('title'), True)
        self.Fields.Note = DatabaseField('Note', self.fieldIndex('note'), True)
        self.Fields.PublicNote = DatabaseField('Public Note', self.fieldIndex('public_note'), True)
        self.Fields.Owner = DatabaseField('Owner', self.fieldIndex('owner'), True)
        self.Fields.Type = DatabaseField('Type', self.fieldIndex('type'), False)
        self.Fields.Workspace = DatabaseField('workspace_id', self.fieldIndex('workspace_id'), False)
        self.Fields.CreationDatetime = DatabaseField('creation_datetime', self.fieldIndex('creation_datetime'), False)
        self.Fields.ModificationDatetime = DatabaseField('modification_datetime', self.fieldIndex('modification_datetime'), False)
        self.Fields.Source = DatabaseField('source', self.fieldIndex('source'), False)
        self.Fields.DocCount = DatabaseField('Doc', self.columnCount() - 2, False)
        self.Fields.Progress = DatabaseField('Progress', self.columnCount() - 1, True)

    def _renameHeaders(self):
        for field in self.Fields.fields():
            self.setHeaderData(field.index,
                               QtCore.Qt.Orientation.Horizontal,
                               field.name,
                               QtCore.Qt.ItemDataRole.DisplayRole)

    def id(self, row):
        return self.data(self.index(row, self.fieldIndex("signage_id")))

    def parent_id(self, row):
        return self.data(self.index(row, self.fieldIndex("parentID")))
    
    def refresh(self):
        self.submitAll()
        self.setFilter(f"workspace_id={AppDatabase.activeWorkspace().id}")
        self.select()
        return super().refresh()
    
    def findIndexById(self, id: int) -> QtCore.QModelIndex|None:
        """Return the QSqlTableModel index of the signage id"""
        for row in range(self.rowCount()):
            index = self.index(row, self.Fields.ID.index)
            if int(self.data(index, QtCore.Qt.ItemDataRole.DisplayRole)) == int(id):
                return index
        return None


class SignageProxyModel(ProxyModel):
    def __init__(self, model):
        super().__init__(model)
        self.setSourceModel(model)
        self.setRecursiveFilteringEnabled(True)

        self.owner_filter = []
        self.evidence_filter = False
        self.evidence_column = 0

    def setOwnerFilter(self, owners: list, column: int):
        self.owner_filter = owners
        self.owner_column = column

    def setEvidenceFilter(self, evidence_only: QtCore.Qt.CheckState, column: int):
        self.evidence_filter = evidence_only
        self.evidence_column = column

    def filterAcceptsRow(self, source_row: int, source_parent: QtCore.QModelIndex) -> bool:
        owner_filter = True
        evidence_filter = True

        # Owner filter
        if len(self.owner_filter) > 0:
            index_owner = self.sourceModel().index(source_row, self.owner_column, source_parent)
            data_owner = self.sourceModel().data(index_owner, QtCore.Qt.ItemDataRole.DisplayRole)
            if data_owner in self.owner_filter:
                owner_filter = True
            else:
                owner_filter = False

        # Evidence filter
        index_evidence = self.sourceModel().index(source_row, self.evidence_column, source_parent)
        data_evidence = self.sourceModel().data(index_evidence, QtCore.Qt.ItemDataRole.DisplayRole)

        # with evidence only
        if self.evidence_filter == QtCore.Qt.CheckState.Checked:
            evidence_filter = bool(data_evidence)
        # without evidence only
        elif self.evidence_filter == QtCore.Qt.CheckState.Unchecked:
            evidence_filter = not bool(data_evidence)
            
        return ProxyModel.filterAcceptsRow(self, source_row, source_parent) and owner_filter and evidence_filter

    def sortTree(self, column=0, order=QtCore.Qt.SortOrder.AscendingOrder, parent=QtCore.QModelIndex()):
        self.sort(column, order)
        for row in range(self.rowCount(parent)):
            child = self.index(row, 0, parent)
            self.sortTree(column, order, child)

class SignageModel(TreeModel):
    """ 
    Tree Model that use QSqlRelationalTableModel to save data in the backend
    
    Dataflow:
    user inserted signage -> SQL Model -> disable SQL Model update in setData -> TreeModel
    user edited signage -> SQL Model -> TreeModel 
    user removed signage -> SQL Model -> TreeModel
    """
    connector_cache = {} # word : [refkey_type, ...], onenote: [object_id]}

    def __init__(self, parent=None):
        super(SignageModel, self).__init__(parent)
        self._source_model = SignageSqlModel()
        self.buildFromSqlModel()
        self.initCache()
        self._sync_enabled = True

    def rootModel(self) -> SignageSqlModel:
        return self._source_model
    
    def buildFromSqlModel(self):
        """Build a tree structure from a flat QSqlRelationalTableModel."""
        self.beginResetModel()

        # Create the root item with headers
        headers = [self._source_model.headerData(i, QtCore.Qt.Orientation.Horizontal) 
                   for i in range(self._source_model.columnCount())]

        self.root_item = TreeItem(headers)

        # Build a dictionary of all items by id
        items_by_id = {}
        for row in range(self._source_model.rowCount()):
            record = self._source_model.record(row)
            data = [record.value(i) for i in range(self._source_model.columnCount())]
            item: TreeItem = TreeItem(data)
            items_by_id[record.value(SignageSqlModel.Fields.ID.index)] = item

        # Link children to parents
        for item_id, item in items_by_id.items():
            record = next(
                (self._source_model.record(r) for r in range(self._source_model.rowCount())
                 if self._source_model.record(r).value(SignageSqlModel.Fields.ID.index) == item_id),
                None
            )
            parent_id = record.value(SignageSqlModel.Fields.ParentID.index) if record else None

            if parent_id is None:
                # Root-level node
                item.parent_item = self.root_item
                self.root_item.child_items.append(item)
            else:
                parent_item: TreeItem = items_by_id.get(parent_id)
                if parent_item:
                    item.parent_item = parent_item
                    parent_item.child_items.append(item)
                else:
                    # Orphaned node (parent missing)
                    item.parent_item = self.root_item
                    self.root_item.child_items.append(item)

        self.endResetModel()

    def setData(self, index: QtCore.QModelIndex, value, role: int) -> bool:
        if role != QtCore.Qt.ItemDataRole.EditRole:
            return False
        
        if self._sync_enabled: # Disabled when inserting
            # --- Apply datachange to sql Model first ---
            signage_id = index.sibling(index.row(), SignageSqlModel.Fields.ID.index).data()
            root_index = self._source_model.findIndexById(signage_id)

            if not root_index:
                return

            if not root_index.isValid():
                return True

            # Ensure we target the same column
            sql_index = root_index.sibling(root_index.row(), index.column())

            # Handle relational field resolution
            relation = self._source_model.relation(index.column())
            if relation.isValid():
                # Convert the display value to the key value
                rel_model = self._source_model.relationModel(index.column())
                key_col = rel_model.fieldIndex(relation.indexColumn())
                display_col = rel_model.fieldIndex(relation.displayColumn())

                match = rel_model.match(
                    rel_model.index(0, display_col), 
                    QtCore.Qt.ItemDataRole.DisplayRole, 
                    value, 
                    hits=1,
                    flags=QtCore.Qt.MatchFlag.MatchExactly
                )
                if match:
                    # Convert found index to key_col index
                    key_value = rel_model.data(match[0].sibling(match[0].row(), key_col),
                                                QtCore.Qt.ItemDataRole.EditRole)
                    result = self._source_model.setData(sql_index, key_value, role)
                else:
                    logger.error(f"⚠️ No matching foreign key for {value}")
                    result = False
            else:
                result = self._source_model.setData(sql_index, value, role)
        else:
            result = True
        
        # --- Apply changes to TreeModel ---
        if result:
            item: TreeItem = self.getItem(index)
            result: bool = item.setData(index.column(), value)
            if not result:
                logger.error(f"Cannot set value:'{value}' to column:'{index.column()}'")

        if result:
            self.dataChanged.emit(index,
                                  index,
                                  [QtCore.Qt.ItemDataRole.DisplayRole,
                                   QtCore.Qt.ItemDataRole.EditRole])

        return result

    def iter_model_rows(self, parent=QtCore.QModelIndex()):
        """
        Recursively yield QModelIndex objects for every row in a tree model.
        Yields one index per row (column 0 by default).
        """
        row_count = self.rowCount(parent)
        for row in range(row_count):
            index = self.index(row, 0, parent)
            if not index.isValid():
                continue

            yield index

            yield from self.iter_model_rows(index)
    
    def initCache(self):
        """Init Connector Cache"""
        self.connector_cache.clear()
        self.connector_cache.setdefault("OneNote", set())
        self.connector_cache.setdefault("Docx", set())

        for row in range(self.rootModel().rowCount()):
            source_json = (self.rootModel().index(row,
                                                    SignageSqlModel.Fields.Source.index)
                                                    .data(QtCore.Qt.ItemDataRole.DisplayRole))
            source: dict = json.loads(source_json)
            application = source.get("application")
            if application == "OneNote":
                object_id = source.get("object_id")
                self.connector_cache.setdefault("OneNote", set()).add(object_id)
            elif application == "Docx":
                object_id = source.get("object_id")
                self.connector_cache.setdefault("Docx", set()).add(object_id)
        
        logger.debug(f"Connector cache's size: {len(self.connector_cache)}")
        

    def insertSignage(self, signage: Signage):
        """Insert new signage into the database"""
        signage.workspace_id = AppDatabase.activeWorkspace().id

        record = self._source_model.record()
        record.setValue(SignageSqlModel.Fields.Refkey.index, signage.refkey)
        record.setValue(SignageSqlModel.Fields.Title.index, signage.title)
        record.setValue(SignageSqlModel.Fields.Owner.index, signage.owner)
        record.setValue(SignageSqlModel.Fields.Type.index, signage.type)
        record.setValue(SignageSqlModel.Fields.Status.index, signage.status)
        record.setValue(SignageSqlModel.Fields.Source.index, signage.source)
        record.setValue(SignageSqlModel.Fields.Note.index, signage.note)
        record.setValue(SignageSqlModel.Fields.PublicNote.index, signage.public_note)
        record.setValue(SignageSqlModel.Fields.ID.index, signage.signage_id)
        record.setValue(SignageSqlModel.Fields.ParentID.index, signage.parentID)
        record.setValue(SignageSqlModel.Fields.Workspace.index, signage.workspace_id)
        record.setValue(SignageSqlModel.Fields.CreationDatetime.index, signage.creation_datetime)
        record.setValue(SignageSqlModel.Fields.ModificationDatetime.index, signage.modification_datetime)

        if not self._source_model.insertRecord(-1, record):
            logger.error(self._source_model.lastError().text())
            return False
        
        self._source_model.refresh()

        # Propagate into TreeModel
        new_id = AppDatabase.lastSignageInserted()

        if not new_id:
            return
                
        new_index = self._source_model.findIndexById(new_id)

        if not new_index:
            logger.error("Index of new record not found!")
            return
        
        new_record = self._source_model.record(new_index.row())

        if signage.parentID:
            parent_index = self.findIndexById(signage.parentID, SignageSqlModel.Fields.ID.index)
            parent_index = parent_index.sibling(parent_index.row(), 0)
        else:
            parent_index = QtCore.QModelIndex()

        row = self.rowCount(parent_index)
        if not self.insertRows(self.rowCount(parent_index), 1, parent_index):
            logger.error("Cannot insert row")
            return False
        
        self._sync_enabled=False
        for column in range(self.columnCount()):
            data = new_record.value(column)
            index: QtCore.QModelIndex = self.index(row, column, parent_index)
            if not self.setData(index, data, QtCore.Qt.ItemDataRole.EditRole):
                logger.error(f"Failed to set data={data}")
                return False
        self._sync_enabled=True
        return True
    
    def deleteRow(self, index: QtCore.QModelIndex) -> bool:
        """Delete a row from the QSqlTableModel and TreeModel and refresh"""
        signage_id = self.data(index.sibling(index.row(),
                                             SignageSqlModel.Fields.ID.index),
                                             QtCore.Qt.ItemDataRole.DisplayRole)
        
        logger.debug(f"Row:{index.row()}, Signage ID:{signage_id}")

        if not signage_id:
            logger.error(f"Signage ID not found from TreeModel: row={index.row()}")
            return False
        
        sql_index = self._source_model.findIndexById(signage_id)
        if not sql_index:
            logger.error(f"Fail to find SQL Model's index for signage ID: {signage_id}")
            return False
        
        if not self._source_model.removeRow(sql_index.row(), QtCore.QModelIndex()):
            logger.error(f"Fail to remove record '{signage_id}' from SQL model:\n\t{self._source_model.lastError().text()}")
            return False

        if not self.removeRows(index.row(), 1, index.parent()):
            logger.error(f"Fail to remove record '{signage_id}' from TreeModel")
            return False    

        self._source_model.refresh()
        return True       

    def summary(self) -> list:
        """Get a summary of signage's status"""
        
        #          | Request | Question | 
        # Open     |   1     |   1      |  
        # Close    |   0     |   1      |  
        # Total    |   2     |   2      |
                
        # Vertical headers
        vheaders = list(AppDatabase.cache_signage_status.strkeys())
        vheaders.append("Total")

        # Horizontal headers
        hheaders = list(AppDatabase.cache_signage_type.strkeys())

        # Init data table
        data = [[0] * len(hheaders) for i in range(len(vheaders))]  

        # Populate data table
        for row in range(self._source_model.rowCount()):
            t_str = (self._source_model.index(row, SignageSqlModel.Fields.Type.index)
                 .data())
            s_str = (self._source_model.index(row, SignageSqlModel.Fields.Status.index)
                 .data())
            
            t = AppDatabase.cache_signage_type.get_int_key(t_str)
            s = AppDatabase.cache_signage_status.get_int_key(s_str)

            if t is not None and s is not None:
                data[int(s)][int(t)] += 1    # row data
                data[-1][int(t)] += 1        # Total row

        return data, vheaders, hheaders

    def updateReviewProgess(self):
        """Update the signage progress bar
        
        Triggered on:
        - Signage refkey update
        - Signage insert
        - Evidence refkey update
        - Evidence insert
        - Evidence delete
        - Evidence status changed
        """
        cache = AppDatabase.queryEvidenceReview()

        self._sync_enabled = False
        for index in self.iter_model_rows():
            refkey = index.sibling(index.row(), SignageSqlModel.Fields.Refkey.index).data(QtCore.Qt.ItemDataRole.DisplayRole)
            signage_type = index.sibling(index.row(), SignageSqlModel.Fields.Type.index).data(QtCore.Qt.ItemDataRole.DisplayRole)
            doc_index =  index.sibling(index.row(), SignageSqlModel.Fields.DocCount.index)
            progress_index = index.sibling(index.row(), SignageSqlModel.Fields.Progress.index)

            if signage_type == "Request" :
                progress = cache.get(refkey, 0)
                if progress:
                    percentage = progress.get('percentage')
                    total = progress.get('total')
                    logger.debug(f'refkey: {refkey}, progress:{percentage}, total:{total}')
                    self.setData(doc_index, total, QtCore.Qt.ItemDataRole.EditRole)
                    self.setData(progress_index, percentage, QtCore.Qt.ItemDataRole.EditRole)
            else:
                self.setData(doc_index, "", QtCore.Qt.ItemDataRole.EditRole)
                self.setData(progress_index, "", QtCore.Qt.ItemDataRole.EditRole)
        self._sync_enabled = True

 