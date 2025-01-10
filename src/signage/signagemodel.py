import logging
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers, PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule

from html2text import html2text

from qtpy import (Qt, QtCore, QtGui, Slot, QtSql)

from models.model import (BaseRelationalTableModel, ProxyModel, DatabaseField)

from db.dbstructure import (Signage, SignageType)
from db.database import AppDatabase

from onenote.onenote_api import (OneNote, Hierarchy, Tag)
from utilities import (utils, config as mconf)

logger = logging.getLogger(__name__)


class SignageTablelModel(BaseRelationalTableModel):

    class Fields:
        ID: DatabaseField
        Status: DatabaseField
        Title: DatabaseField
        RefKey: DatabaseField
        Note: DatabaseField
        Type: DatabaseField
        Workspace: DatabaseField
        Uid: DatabaseField
        CreationDatetime: DatabaseField
        ModificationDatetime: DatabaseField
        Link: DatabaseField
        Owner: DatabaseField
        Evidence: DatabaseField
        EvidenceEOL: DatabaseField
        PublicNote: DatabaseField

    def __init__(self):
        super().__init__()
        self.setEditStrategy(QtSql.QSqlTableModel.EditStrategy.OnFieldChange)
        self.status_color_cache = {}
        self.onenote: OneNote = None

        self.setTable("signage")
        self.init_fields()

        self.setRelation(self.Fields.Status.index,
                         QtSql.QSqlRelation("signage_status", "status_id", "Status"))
        self.setRelation(self.Fields.Type.index,
                         QtSql.QSqlRelation("signage_type", "type_id", "type"))

        self.setFilter(f"workspace_id={AppDatabase.active_workspace.id}")

        self.setHeaderData(self.Fields.RefKey.index,
                           Qt.Orientation.Horizontal,
                           "refKey")
        self.setHeaderData(self.Fields.Title.index,
                           Qt.Orientation.Horizontal,
                           "Title")
        self.setHeaderData(self.Fields.Note.index,
                           Qt.Orientation.Horizontal,
                           "Note")
        self.setHeaderData(self.Fields.Status.index,
                           Qt.Orientation.Horizontal,
                           "Status")
        self.setHeaderData(self.Fields.Type.index,
                           Qt.Orientation.Horizontal,
                           "Type")
        self.setHeaderData(self.Fields.Owner.index,
                           Qt.Orientation.Horizontal,
                           "Owner")
        self.setHeaderData(self.Fields.Evidence.index,
                           Qt.Orientation.Horizontal,
                           "")
        self.setHeaderData(self.Fields.EvidenceEOL.index,
                           Qt.Orientation.Horizontal,
                           "Review")
        self.select()

    def init_fields(self):
        self.Fields.ID = DatabaseField('signage_id', self.fieldIndex('signage_id'), False)
        self.Fields.Status = DatabaseField('status_id', self.fieldIndex('status_id'), True)
        self.Fields.RefKey = DatabaseField('refKey', self.fieldIndex('refKey'), True)
        self.Fields.Title = DatabaseField('title', self.fieldIndex('title'), True)
        self.Fields.Note = DatabaseField('note', self.fieldIndex('note'), True)
        self.Fields.Owner = DatabaseField('owner', self.fieldIndex('owner'), True)
        self.Fields.Type = DatabaseField('type_id', self.fieldIndex('type_id'), True)
        self.Fields.Workspace = DatabaseField('workspace_id', self.fieldIndex('workspace_id'), False)
        self.Fields.Uid = DatabaseField('uid', self.fieldIndex('uid'), False)
        self.Fields.CreationDatetime = DatabaseField('creation_datetime', self.fieldIndex('creation_datetime'), False)
        self.Fields.ModificationDatetime = DatabaseField('modification_datetime', self.fieldIndex('modification_datetime'), False)
        self.Fields.Link = DatabaseField('link', self.fieldIndex('link'), False)
        self.Fields.Evidence = DatabaseField('evidence', self.fieldIndex('evidence'), True)
        self.Fields.EvidenceEOL = DatabaseField('evidence_eol', self.fieldIndex('evidence_eol'), True)
        self.Fields.PublicNote = DatabaseField('public_note', self.fieldIndex('public_note'), False)

    def querySignage(self, row) -> Signage:
        signage = Signage(note=self.index(row, self.Fields.Note.index).data(Qt.ItemDataRole.DisplayRole),
                          public_note=self.index(row, self.Fields.PublicNote.index).data(Qt.ItemDataRole.DisplayRole),
                          status_id=self.index(row, self.Fields.Status.index).data(Qt.ItemDataRole.DisplayRole),
                          owner=self.index(row, self.Fields.Owner.index).data(Qt.ItemDataRole.DisplayRole),
                          type_id=self.index(row, self.Fields.Type.index).data(Qt.ItemDataRole.DisplayRole),
                          refKey=self.index(row, self.Fields.RefKey.index).data(Qt.ItemDataRole.DisplayRole),
                          title=self.index(row, self.Fields.Title.index).data(Qt.ItemDataRole.DisplayRole),
                          creation_datetime=self.index(row, self.Fields.CreationDatetime.index).data(Qt.ItemDataRole.DisplayRole),
                          modification_datetime=self.index(row, self.Fields.ModificationDatetime.index).data(Qt.ItemDataRole.DisplayRole),
                          link=self.index(row, self.Fields.Link.index).data(Qt.ItemDataRole.DisplayRole),
                          signage_id=self.index(row, self.Fields.ID.index).data(Qt.ItemDataRole.DisplayRole),
                          workspace_id=self.index(row, self.Fields.Workspace.index).data(Qt.ItemDataRole.DisplayRole),
                          uid=self.index(row, self.Fields.Uid.index).data(Qt.ItemDataRole.DisplayRole))
        return signage

    def get_last_id(self, signage_type: str, prefix: str) -> str:
        last_id = AppDatabase.queryRefKey(signage_type, prefix)
        return last_id

    def data(self, index: QtCore.QModelIndex, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.ForegroundRole:
            status_idx = self.index(index.row(), self.Fields.Status.index)  # Column with status text
            status = self.data(status_idx, Qt.ItemDataRole.DisplayRole)

            if status:
                if status not in self.status_color_cache:
                    self.status_color_cache[status] = self.fetch_status_color(status)

                color_hex = self.status_color_cache[status]
                if color_hex:
                    return QtGui.QColor(color_hex)

        return super().data(index, role)

    def getLink(self, row: int) -> str:
        link_idx = self.index(row, self.Fields.Link.index)
        link = self.data(link_idx, Qt.ItemDataRole.DisplayRole)
        return link

    def fetch_status_color(self, status):
        query = QtSql.QSqlQuery()
        query.prepare("SELECT color FROM signage_status WHERE status = ?")
        query.addBindValue(status)
        if query.exec() and query.next():
            return query.value(0)
        return None

    def fetch_onenote(self):
        section_id = AppDatabase.active_workspace.OESectionID()
        if section_id != "":

            if self.onenote is None:
                self.onenote, err = OneNote.connect()

                if err is not None:
                    return

            page_str, err = self.onenote.get_hierarchy(section_id, 4)

            if err is not None:
                return

            element_tree = ET.fromstring(page_str)
            hierarchy = Hierarchy(element_tree)
            all_tags: list[Tag] = []

            regex = mconf.default_regex if mconf.settings.value("regex") is None or mconf.settings.value("regex") == "" else mconf.settings.value("regex")

            for page in hierarchy:
                page_tree = ET.fromstring(self.onenote.get_page_content(page_id=page.id))
                tags = self.onenote.get_tags(page.id, page_tree)
                all_tags.extend(tags)

            self.beginInsertRows(QtCore.QModelIndex(), self.rowCount(), self.rowCount() + len(all_tags))
            for tag in all_tags:
                record = self.record()
                record.setValue(self.Fields.Evidence.index, 0)
                record.setValue(self.Fields.EvidenceEOL.index, 0)
                record.setValue(self.Fields.RefKey.index, utils.find_match(tag.text, regex))
                record.setValue(self.Fields.Title.index, tag.text)
                record.setValue(self.Fields.Status.index, 1)
                if tag.type in AppDatabase.cache_signage_type.keys():
                    signage_type: SignageType = AppDatabase.cache_signage_type.get(tag.type)
                    signage_type_id = int(signage_type.type_id)
                else:
                    signage_type_id = 1
                record.setValue(self.Fields.Type.index, signage_type_id)
                record.setValue(self.Fields.Note.index, "")
                record.setValue(self.Fields.Workspace.index, AppDatabase.active_workspace.id)
                record.setValue(self.Fields.Uid.index, tag.object_id)
                record.setValue(self.Fields.Link.index, tag.link)
                record.setValue(self.Fields.CreationDatetime.index, tag.creationTime)
                record.setValue(self.Fields.ModificationDatetime.index, tag.lastModifiedTime)

                self.insertRecord(self.rowCount(), record)

            self.endInsertRows()
            self.submitAll()
            self.refresh()

    def signageRecord(self, signage: Signage) -> QtSql.QSqlRecord:
        """Create a QSqlRecord by mapping the model object to database object"""
        record = self.record()
        record.setValue(self.Fields.Evidence.index, 0)
        record.setValue(self.Fields.EvidenceEOL.index, 0)
        record.setValue(self.Fields.RefKey.index, signage.refKey)
        record.setValue(self.Fields.Title.index, signage.title)
        record.setValue(self.Fields.Status.index, signage.status_id)
        record.setValue(self.Fields.Type.index, signage.type_id)
        record.setValue(self.Fields.Note.index, signage.note)
        record.setValue(self.Fields.PublicNote.index, signage.public_note)
        record.setValue(self.Fields.Workspace.index, signage.workspace_id)
        record.setValue(self.Fields.Owner.index, signage.owner)
        record.setValue(self.Fields.Link.index, signage.link)

        return record

    def updateStatus(self, rows: list[int], status: int):
        for row in rows:
            res = self.setData(self.index(row, self.Fields.Status.index), status, Qt.ItemDataRole.EditRole)

            if res:
                self.submit()

        self.refresh()

    def updateOwner(self, rows: list[int], owner: str):
        for row in rows:
            res = self.setData(self.index(row, self.Fields.Owner.index), owner, Qt.ItemDataRole.EditRole)

            if res:
                self.submit()

        self.refresh()

    def updateProgress(self):
        self.review_progress = AppDatabase.reviewProgress()
        for row in range(self.rowCount()):
            refkey = self.data(self.index(row, self.Fields.RefKey.index), Qt.ItemDataRole.EditRole)
            item = self.review_progress.get(refkey)
            if item:
                progression = f'{item["progress"]},({item["closed"]}/{item["count"]})'

                res = self.setData(self.index(row, self.Fields.ReviewProgress.index), progression, Qt.ItemDataRole.EditRole)

                if res:
                    self.submit()

        self.refresh()

    def insertSignage(self, signage: Signage) -> bool:
        record = self.signageRecord(signage)

        self.layoutAboutToBeChanged.emit()
        self.beginInsertRows(QtCore.QModelIndex(), self.rowCount(), self.rowCount())
        inserted = self.insertRecord(-1, record)
        self.endInsertRows()
        self.layoutChanged.emit()

        if inserted == True:
            self.refresh()
            self.submitAll()
        else:
            err = self.lastError().text()
            logger.error(f"Cannot insert: {signage} - Error: {err}")

        return inserted

    def export2Excel(self, signage_type: list, destination: str, public_note: bool = False) -> Exception | None:
        wb = Workbook()
        ws = wb.active
        ws.title = "main"

        if public_note:
            headers = ["RefKey", "Title", "Owner", "Type", "Evidence", "Note"]
            xrange = "A1:F"
            model_fields = [SignageTablelModel.Fields.RefKey.index,
                            SignageTablelModel.Fields.Title.index,
                            SignageTablelModel.Fields.Owner.index,
                            SignageTablelModel.Fields.Type.index,
                            SignageTablelModel.Fields.Evidence.index,
                            SignageTablelModel.Fields.PublicNote.index]
        else:
            headers = ["RefKey", "Title", "Owner", "Type", "Evidence"]
            xrange = "A1:E"
            model_fields = [SignageTablelModel.Fields.RefKey.index,
                            SignageTablelModel.Fields.Title.index,
                            SignageTablelModel.Fields.Owner.index,
                            SignageTablelModel.Fields.Type.index,
                            SignageTablelModel.Fields.Evidence.index]

        for column in range(len(headers)):
            ws.cell(row=1, column=column + 1, value=headers[column])

        record_count = 1
        for row in range(self.rowCount()):
            record = self.record(row)
            if record.value(self.Fields.Type.index) in signage_type:
                record_count += 1
                values = list(map(record.value, model_fields))
   
                if public_note:
                    if values[-1] is not None:
                        values[-1] = html2text(values[-1])

                ws.append(values)

        if record_count == 1:
            record_count = 2

        table = Table(displayName="Table1", ref=f"{xrange}{record_count}")
        style = TableStyleInfo(name="TableStyleMedium2",
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=False)
        table.tableStyleInfo = style

        ws.column_dimensions["B"].width = 150.0
        ws.column_dimensions["A"].number_format = numbers.FORMAT_TEXT

        # Conditional formatting
        red_text = Font(color="9C0006")
        red_fill = PatternFill(bgColor="FFC7CE")
        ws.conditional_formatting.add(f"E2:E{record_count}", CellIsRule(operator='equal', formula=['0'], stopIfTrue=False, font=red_text, fill=red_fill))

        for cell in ws['E':'E']:
            cell.alignment = Alignment(horizontal="center")

        try:
            ws.add_table(table)
            wb.save(destination)
        except Exception as e:
            logger.error(f"Cannot save the file - Error:{e}")
            return e

    def deleteRows(self, indexes: list[QtCore.QModelIndex]) -> bool:
        """Delete a row from the model and refresh the model"""

        self.beginRemoveRows(QtCore.QModelIndex(), indexes[0].row(), indexes[-1].row())

        for index in indexes:
            self.removeRow(index.row())

        self.endRemoveRows()

        self.refresh()


    def importFromExcels(self, files: list, update: bool = True):
        df = utils.mergeExcelFiles(files, drop_duplicate="first", outfile="")
        # df.fillna(' ', method='ffill', inplace=True)
        df.sort_values(by=['RefKey'], inplace=True)

        self.sort(self.Fields.RefKey.index, Qt.SortOrder.AscendingOrder)
        for index, df_row in df.iterrows():
            df_refkey = f'{df_row["RefKey"]:>{'0'}{3}}' # format rekey with leading '0'
            df_type = df_row["Type"]

            for row in range(self.rowCount()):
                refkey = self.data(self.index(row, self.Fields.RefKey.index), Qt.ItemDataRole.EditRole)
                signagetype = self.data(self.index(row, self.Fields.Type.index), Qt.ItemDataRole.EditRole)

                if df_refkey == refkey and df_type == signagetype:
                    if update:
                        res = self.setData(self.index(row, self.Fields.Title.index), df_row["Title"], Qt.ItemDataRole.EditRole)
                        if res:
                            self.submit()
                    break
            else:  # Insert if not found
                if df_refkey != "":
                    signage_type: SignageType = AppDatabase.cache_signage_type.get(df_type)

                    if signage_type is None:
                        signage_type = AppDatabase.cache_signage_type.get('Request') # apply "Request" as default type to avoid fail import

                    try:
                        signage = Signage(status_id=1,
                                          owner=df_row["Owner"],
                                          type_id=signage_type.type_id,
                                          refKey=df_refkey,
                                          title=df_row["Title"],
                                          workspace_id=AppDatabase.active_workspace.id)
                        self.insertSignage(signage)
                    except Exception as e:
                        logger.error(f"importFromExcels > Error importing Signage: refkey={df_refkey} -- Error: {e}")


        self.refresh()


class SignageProxyModel(ProxyModel):
    def __init__(self, model):
        super().__init__(model)

        self.setSourceModel(model)

        self.owner_filter = []
        self.evidence_filter = False
        self.evidence_column = 0

    def setOwnerFilter(self, owners: list, column: int):
        self.owner_filter = owners
        self.owner_column = column

    def setEvidenceFilter(self, evidence_only: QtCore.Qt.CheckState, column: int):
        self.evidence_filter = evidence_only
        self.evidence_column = column

    def filterAcceptsRow(self, source_row: int, source_parent: QtCore.QModelIndex) -> bool:
        owner_filter = True
        evidence_filter = True

        # Owner filter
        if len(self.owner_filter) > 0:
            index_owner = self.sourceModel().index(source_row, self.owner_column, source_parent)
            data_owner = self.sourceModel().data(index_owner)
            if data_owner in self.owner_filter:
                owner_filter = True
            else:
                owner_filter = False

        # Evidence filter
        index_evidence = self.sourceModel().index(source_row, self.evidence_column, source_parent)
        data_evidence = self.sourceModel().data(index_evidence)
        # with evidence only
        if self.evidence_filter == QtCore.Qt.CheckState.Checked:
            evidence_filter = bool(data_evidence)
        # without evidence only
        elif self.evidence_filter == QtCore.Qt.CheckState.Unchecked:
            evidence_filter = not bool(data_evidence)
            
        return ProxyModel.filterAcceptsRow(self, source_row, source_parent) and owner_filter and evidence_filter