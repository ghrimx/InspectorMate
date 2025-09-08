# Standard library imports.
import logging
from pandas import DataFrame
from datetime import datetime, timezone

# Related third party imports.
from qtpy import (Qt, QtCore, QtGui, QtSql)
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers, PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from html2text import html2text

# Local application/library specific imports.
from database.database import AppDatabase, Signage, SignageType, SignageStatus, Cache
from models.model import DatabaseField, ProxyModel
from models.treemodel import (TreeItem, TreeModel)

from utilities import config as mconf
from utilities.utils import mergeExcelFiles, find_match

from onenote.msonenote import OnenoteModel, OE

from qt_theme_manager import theme_icon_manager, Theme

logger = logging.getLogger(__name__)


class SignageTreeModel(TreeModel):
    cache_oe_signage = []

    class Fields:
        ID: DatabaseField
        ParentID: DatabaseField
        Status: DatabaseField
        Title: DatabaseField
        Refkey: DatabaseField
        Note: DatabaseField
        Type: DatabaseField
        Workspace: DatabaseField
        CreationDatetime: DatabaseField
        ModificationDatetime: DatabaseField
        Source: DatabaseField
        Owner: DatabaseField
        Evidence: DatabaseField
        EvidenceEOL: DatabaseField
        PublicNote: DatabaseField

        @classmethod
        def fields(cls) -> list[DatabaseField]:
            """Return the list of fields.
                
            The order of item in the list matters and should match the fields order from the database.
            """
            return [
                    cls.Refkey,
                    cls.Title,
                    cls.Owner,
                    cls.Type,
                    cls.Status,
                    cls.Source,
                    cls.Note,
                    cls.PublicNote,
                    cls.Evidence,
                    cls.EvidenceEOL,
                    cls.CreationDatetime,
                    cls.ModificationDatetime,
                    cls.ID,
                    cls.ParentID,
                    cls.Workspace
                    ]

    def __init__(self, parent = None):
        super(SignageTreeModel, self).__init__(parent)
        self.headers = []
        self.root_item: TreeItem = None
        self._data = None

        self.initSourceModel()
        self.initFields()  
        self.loadData()

        self.summary()
        self.cacheOESignage()

    def data(self, index: QtCore.QModelIndex, role=QtCore.Qt.ItemDataRole.DisplayRole):

        # Change foreground color of signage table's rows depending on the signage status
        if role == QtCore.Qt.ItemDataRole.ForegroundRole:
            status_idx = index.sibling(index.row(), self.Fields.Status.index)  # Column with status text
            status = self.data(status_idx, QtCore.Qt.ItemDataRole.DisplayRole)

            signage_status = self.cacheSignageStatus().get(status)
            if signage_status is not None :
                
                color_hex = signage_status.color

                if color_hex == "#000000" and theme_icon_manager.get_theme() == Theme.DARK:
                    color_hex = theme_icon_manager.get_theme_color().name(QtGui.QColor.NameFormat.HexRgb)

                if color_hex:
                    return QtGui.QColor(color_hex)

        return super().data(index, role)
    
    def rootItem(self) -> TreeItem:
        return self.root_item

    def initSourceModel(self):
        self._source_model = QtSql.QSqlTableModel()
        self._source_model.setTable("signage")
        self._source_model.setFilter(f"workspace_id={AppDatabase.activeWorkspace().id}")
        
        success = self._source_model.select()

        if not success:
            logger.error(f"initSourceModel: Failed to init the source model - ERROR: {self._source_model.lastError().text()}")
            return

        for column in range(self._source_model.columnCount()):
            self.headers.append(self._source_model.record(0).fieldName(column))

    def refreshSourceModel(self) -> tuple[bool, str]:
        """Refresh the QSqlTableModel"""
        if not self._source_model.select():
            err = f"refreshSourceModel > Failed to refresh the source model - ERROR: {self._source_model.lastError().text()}"
            logger.error(err)
            return False, err
        
        self._source_model.setFilter(f"workspace_id={AppDatabase.activeWorkspace().id}")

        return True, "Source model refreshed successfully!"
    
    def refreshTreeModel(self):
        """Refresh/ReLoad Data of TreeModel"""
        self.beginResetModel()
        ok, err = self.loadData()
        self.endResetModel()

        return ok, err

    def initFields(self):
        self.Fields.ID = DatabaseField('signage_id', self.fieldIndex('signage_id'), False)
        self.Fields.ParentID = DatabaseField('parentID', self.fieldIndex('parentID'), False)
        self.Fields.Status = DatabaseField('Status', self.fieldIndex('status'), True)
        self.Fields.Refkey = DatabaseField('Refkey', self.fieldIndex('refkey'), True)
        self.Fields.Title = DatabaseField('Title', self.fieldIndex('title'), True)
        self.Fields.Note = DatabaseField('Note', self.fieldIndex('note'), True)
        self.Fields.PublicNote = DatabaseField('Public Note', self.fieldIndex('public_note'), True)
        self.Fields.Owner = DatabaseField('Owner', self.fieldIndex('owner'), True)
        self.Fields.Type = DatabaseField('Type', self.fieldIndex('type'), True)
        self.Fields.Workspace = DatabaseField('workspace_id', self.fieldIndex('workspace_id'), False)
        self.Fields.CreationDatetime = DatabaseField('creation_datetime', self.fieldIndex('creation_datetime'), False)
        self.Fields.ModificationDatetime = DatabaseField('modification_datetime', self.fieldIndex('modification_datetime'), False)
        self.Fields.Source = DatabaseField('source', self.fieldIndex('source'), False)
        self.Fields.Evidence = DatabaseField('Evidence', self.fieldIndex('evidence'), False)
        self.Fields.EvidenceEOL = DatabaseField('Review', self.fieldIndex('evidence_eol'), True)
            
    def sourceModel(self):
        return self._source_model
    
    def fieldIndex(self, name: str) -> int:
        return self._source_model.fieldIndex(name)

    def traverse_node(self, node: TreeItem, value: str|int) -> TreeItem:
        """Traverse the SignageTreeModel to find the parent signage"""
        stack = [node]  
        
        while stack:
            current: TreeItem = stack.pop() 
            
            for i in range(current.childCount()):
                if current.child(i).data(self.Fields.ID.index) == value:
                    return current.child(i)

                if current.child(i).childCount() > 0:  
                    stack.append(current.child(i))

    def loadData(self):
        """Load the SignageTreeModel with data from the database via the QSqlTableModel"""

        headers = []
        for field in self.Fields.fields():
            headers.append(field.name)

        self.root_item = TreeItem(headers.copy())

        for row in range(self._source_model.rowCount()):

            parent_id = self._source_model.record(row).value(self.Fields.ParentID.index)

            # create child record if parentID is not empty
            if parent_id != "":
                parent_node = self.traverse_node(self.root_item, parent_id)
                if parent_node is not None:
                    parent_node.insertChildren(parent_node.childCount(), 1, self.root_item.columnCount())
            else:
                parent_node: TreeItem = self.root_item
                col_count = self.root_item.columnCount()
                parent_node.insertChildren(parent_node.childCount(), 1, col_count)

            # set data to the Treeitem
            for column in range(self._source_model.columnCount()):
                try:
                    child_node: TreeItem = parent_node.lastChild()
                    child_node.setData(column, self._source_model.record(row).value(column))
                except Exception as e:
                    err = f"loadData: Cannot set data from model's field={headers[column]} ({row}:{column})\n\tdata={self._source_model.record(row).value(column)}\n\tERROR: {e}"
                    logger.error(err)
                    return False, err
        
        return True, "Data Loaded!"

    def insertSignage(self, signage: Signage,
                      parent: QtCore.QModelIndex = QtCore.QModelIndex,
                      update_treemodel: bool = True) -> tuple[bool, str]:
        """Insert new signage into the database"""
        signage.workspace_id = AppDatabase.activeWorkspace().id

        record = self.sourceModel().record()
        record.setValue(self.Fields.Refkey.index, signage.refkey)
        record.setValue(self.Fields.Title.index, signage.title)
        record.setValue(self.Fields.Owner.index, signage.owner)
        record.setValue(self.Fields.Type.index, signage.type)
        record.setValue(self.Fields.Status.index, signage.status)
        record.setValue(self.Fields.Source.index, signage.source)
        record.setValue(self.Fields.Note.index, signage.note)
        record.setValue(self.Fields.PublicNote.index, signage.public_note)
        record.setValue(self.Fields.Evidence.index, signage.evidence)
        record.setValue(self.Fields.EvidenceEOL.index, signage.evidence_eol)
        record.setValue(self.Fields.ID.index, signage.signage_id)
        record.setValue(self.Fields.ParentID.index, signage.parentID)
        record.setValue(self.Fields.Workspace.index, signage.workspace_id)
        record.setValue(self.Fields.CreationDatetime.index, signage.creation_datetime)
        record.setValue(self.Fields.ModificationDatetime.index, signage.modification_datetime)

        self.sourceModel().beginInsertRows(QtCore.QModelIndex(), self.sourceModel().rowCount(), self.sourceModel().rowCount())
        inserted = self.sourceModel().insertRecord(-1, record)
        self.sourceModel().endInsertRows()
        
        if not inserted:
            err = self.sourceModel().lastError().text()
            logger.error(f"Cannot insert record into the database: {signage} - Error:{err}")
            return False, err

        if not self.sourceModel().submitAll():
            err = self.sourceModel().lastError().text()
            logger.error(f"Cannot submit changes to the database: {signage} - Error: {err}")
            return False, err
        
        ok, err = self.refreshSourceModel()
        if not ok:
            return False, err

        if update_treemodel:
            # Fetch the last record inserted
            last_record = self._source_model.record(self.sourceModel().rowCount() - 1)
            if last_record.isNull(self.Fields.ID.index):
                err = "Record signage_id is Null"
                logger.error(f"{err}: {signage}")
                return False, err

            ok, err = self.insertRecordIntoTreeModel(last_record, parent)

        return ok, err

    def insertRecordIntoTreeModel(self, record: QtSql.QSqlRecord, parent: QtCore.QModelIndex) -> tuple[bool, str]:
        """Insert the db record signage into the model"""

        row = self.rowCount(parent)
 
        if not self.insertRows(self.rowCount(parent), 1, parent):
            err = "Cannot insert row"
            logger.error(err)
            return False, err

        for column in range(self.columnCount()):
            data = record.value(column)
            child: QtCore.QModelIndex = self.index(row, column, parent)

            if not self.setData(child, data, QtCore.Qt.ItemDataRole.EditRole):
                err = f"Failed to set data={data}"
                logger.error(err)
                return False, err       
        
        return True, "Record inserted successfully!"

    # TODO
    def selectRequests(self):
        query = QtSql.QSqlQuery()
        query.prepare("""
                      SELECT
                        refkey,
                        title,
                        id,
                        parentID
                      FROM signage
                      WHERE 
                        signage.workspace_id = :workspace_id
                      AND
                        type = :type
                      """)
        query.bindValue(":workspace_id", AppDatabase.activeWorkspace().id)
        query.bindValue(":type_id", AppDatabase.cache_signage_type['Request'].type_id)

        query.exec()

        request = []

        if not query.exec():
            logger.error(f"selectRequest: Query execution failed - ERROR: {query.lastError().text()}")
        else:
            while query.next():
                request = Signage(refkey=query.value(0), title=query.value(1), signage_id=query.value(2))
            return request
        
    def getSignageLastRefkey(self, signage_type: str = "", pattern: str = "") -> str:
        """Query the last signage refkey from the database"""
        
        query = QtSql.QSqlQuery()
        query.prepare("""
                        SELECT MAX(refkey)
                        FROM
                            signage
                        WHERE
                            signage.workspace_id = :workspace_id
                        AND
                            signage.refkey REGEXP :pattern
                        AND
                            signage.type = (SELECT uid FROM signage_type WHERE name = :signage_type)
                        ORDER BY
                            signage.refkey;""")

        query.bindValue(":workspace_id", AppDatabase.activeWorkspace().id)
        query.bindValue(":pattern", pattern)
        query.bindValue(":signage_type", signage_type)

        query.exec()

        last_refkey = ""

        if not query.exec():
            logger.error(f"getSignageLastRefkey: Query execution failed with error : {query.lastError().text()}")
        elif query.next():
            last_refkey = query.value(0)
        else:
            logger.error(f"getSignageLastRefkey: No rows found with query : {query.lastQuery()} {AppDatabase.activeWorkspace().id}")

        return last_refkey
    
    # TEST
    def getLastInsertSignageId(self) -> str:
        query = QtSql.QSqlQuery()
        query.exec("SELECT last_insert_rowid();")  # SQLite-specific query for last inserted ID
        last_inserted_id = ""

        if query.next():
            last_inserted_id = query.value(0)
        return last_inserted_id

    def cacheSignageType(self) -> Cache:
        """Wrapper on AppDatabase method cache_signage_type
        return: dict
            key : type_id
            value : class SignageType
        """
        return AppDatabase.cache_signage_type
    
    def cacheSignageStatus(self) -> Cache:
        """Wrapper on AppDatabase method cache_signage_status
        return: dict
            key : status_id
            value : class SignageStatus
        """
        return AppDatabase.cache_signage_status
    
    def cacheSignageOwners(self) -> list:
        return mconf.settings.value("owners", [], "QStringList")
    
    def activeWorkspace(self):
        return AppDatabase.activeWorkspace()
    
    def updateReviewProgess(self) -> tuple[bool, str]:
        """Update review progress bar"""
        ok, err = self.refreshSourceModel()
        
        if not ok:
            return False, err
        
        for row in range(self._source_model.rowCount()):
            id = self._source_model.record(row).value(self.Fields.ID.index)
            count_of_evidence = self._source_model.record(row).value(self.Fields.Evidence.index)
            count_of_eol_evidence = self._source_model.record(row).value(self.Fields.EvidenceEOL.index)

            node = self.traverse_node(self.root_item, id)

            if node:
                try:
                    node.setData(self.Fields.Evidence.index, count_of_evidence)
                    node.setData(self.Fields.EvidenceEOL.index, count_of_eol_evidence)
                except Exception as e:
                    err = f"updateReviewProgess > Cannot set data of signage_id={id}\n\tERROR: {e}"
                    logger.error(err)
                    return False, err
        
        return True, "Progress review updated successfully!"

    def deleteRow(self, treemodel_index: QtCore.QModelIndex) -> tuple[bool, str]:
        """Delete a row from the QSqlTableModel and TreeModel and refresh"""

        if not treemodel_index.isValid():
            msg = "Invalid index provided."
            logger.error(f"deleteRow > {msg}")
            return False, msg

        id = self.data(treemodel_index.sibling(treemodel_index.row(), self.Fields.ID.index),
                       QtCore.Qt.ItemDataRole.DisplayRole)
        source = self.data(treemodel_index.sibling(treemodel_index.row(), self.Fields.Source.index),
                       QtCore.Qt.ItemDataRole.DisplayRole)
        
        oe_id = source.split(',')[-1].split(':')[-1][1:-2].strip()
        if oe_id in self.cache_oe_signage:
            self.cache_oe_signage.remove(oe_id)

        source_index = self.getSourceIndexById(id)

        if source_index is None:
            msg = "QSqlTableModel's index is None."
            logger.error(f"deleteRow > {msg} : signage_id={id}")
            return False, msg

        self.sourceModel().beginRemoveRows(source_index.parent(), source_index.row(), source_index.row() + 1)
        if not self.sourceModel().removeRow(source_index.row(), source_index.parent()):
            self.sourceModel().endRemoveRows()
            msg = "Cannot remove row from QSqlTableModel."
            logger.error(f"deleteRow > {msg} : signage_id={id}")
            return False, msg
        self.sourceModel().endRemoveRows()

        if not self.sourceModel().submitAll():
            msg = "Cannot submit changes to QSqlTableModel."
            logger.error(f"deleteRow > {msg} : signage_id={id}\nError: {self.sourceModel().lastError().text()}")
            return False, msg

        ok, err = self.refreshSourceModel()
        if not ok:
            msg = "Cannot refresh source model."
            logger.error(f"deleteRow > {err} : signage_id={id}")
            return False, msg
        
        if not self.removeRows(treemodel_index.row(), 1, treemodel_index.parent()):
            msg = "Cannot remove row from SignageTreeModel."
            logger.error(f"deleteRow > {msg} : signage_id={id}")
            return False, msg

        return True, "Row deleted successfully."

    def getSourceIndexById(self, id: int) -> QtCore.QModelIndex|None:
        """Return the QSqlTableModel index of the signage id"""
        source_index = None
        for row in range(self.sourceModel().rowCount()):
            index = self.sourceModel().index(row, self.Fields.ID.index)
            if int(self.sourceModel().data(index, role=QtCore.Qt.ItemDataRole.DisplayRole)) == int(id):
                source_index = index
        return source_index
    
    def traverse_index(self, index: QtCore.QModelIndex, value: int) -> QtCore.QModelIndex|None:
        stack = [index]
        while stack:
            current: QtCore.QModelIndex = stack.pop()
            for i in range(self.rowCount(current)):
                child = self.index(i, current.column(), current)
                child_uid = child.sibling(i, self.Fields.ID.index).data(QtCore.Qt.ItemDataRole.DisplayRole)
                logger.debug(f"child_uid={child_uid}")
                if int(child_uid) == value:
                    return child
                if self.rowCount(child) > 0:
                    stack.append(child)
    
    def getTreeModelIndexById(self, id: int) -> QtCore.QModelIndex|None:
        """Return the TreeModel index of the signage id"""
        treemodel_index = None

        logger.debug(f"rowCount={range(self.rowCount(self.parent()))}")
        logger.debug(f"Search Signage={id}, type={type(id)}")

        for row in range(self.rowCount(self.parent())):
            index = self.index(row, 0, self.parent())

            if not index.isValid():
                return None
            
            uid = index.sibling(index.row(), self.Fields.ID.index).data(QtCore.Qt.ItemDataRole.DisplayRole)
            logger.debug(f"current uid={uid}")

            if int(uid) == id:
                treemodel_index = index
                break
            elif self.rowCount(index) > 0:
                treemodel_index = self.traverse_index(index, id)
            else:
                continue

        return treemodel_index
    
    def updateField(self, treemodel_index: QtCore.QModelIndex, column: int, new_value: int|str, update_treemodel: bool = True):
        """Update the source model field"""

        if not treemodel_index.isValid():
            return False, "updateField > Invalid index"

        # Get id of the signage to update
        id = treemodel_index.sibling(treemodel_index.row(), self.Fields.ID.index).data(QtCore.Qt.ItemDataRole.DisplayRole)

        source_index = self.getSourceIndexById(id)

        if source_index is None:
            msg = "QSqlTableModel's index is None."
            logger.error(f"updateField > {msg} : signage_id={id}")
            return False, msg

        # Update QSqlTableModel
        if not self.sourceModel().setData(source_index.sibling(source_index.row(), column),
                                          new_value,
                                          QtCore.Qt.ItemDataRole.EditRole):
            err = f"updateField > Source model data cannot be updated : signage_id={id} - Error:{self.sourceModel().lastError().text()}"
            logger.error(err)
            return False, err
        
        # Submit changes to db
        if not self.sourceModel().submitAll():
            err = self.sourceModel().lastError().text()
            logger.error(f"updateField > Cannot submit changes to the database: signage_id={id} - Error: {err}")
            return False, err
        
        # Refresh QSqlTableModel
        ok, err = self.refreshSourceModel()
        if not ok:
            return False, err
        
        # Update TreeModel
        if update_treemodel:
            if not self.setData(treemodel_index.sibling(treemodel_index.row(), column),
                                new_value,
                                QtCore.Qt.ItemDataRole.EditRole):
                err = f"updateField > Unable to update TreeModel data: signage_id={id}"
                logger.error(err)
                return False, err
        
        return True, "updateField > Field updated successfully!"
    
    def export2Excel(self, types: list, statuses: list, destination: str, include_publicnote: bool = False):
        """Export Signage to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "main"

        if include_publicnote:
            headers = ["Refkey", "Title", "Status", "Owner", "Type", "Evidence", "Note"]
            xrange = "A1:G"
            model_fields = [self.Fields.Refkey.index,
                            self.Fields.Title.index,
                            self.Fields.Status.index,
                            self.Fields.Owner.index,
                            self.Fields.Type.index,
                            self.Fields.Evidence.index,
                            self.Fields.PublicNote.index]
        else:
            headers = ["Refkey", "Title", "Status", "Owner", "Type", "Evidence"]
            xrange = "A1:F"
            model_fields = [self.Fields.Refkey.index,
                            self.Fields.Title.index,
                            self.Fields.Status.index,
                            self.Fields.Owner.index,
                            self.Fields.Type.index,
                            self.Fields.Evidence.index]

        for column in range(len(headers)):
            ws.cell(row=1, column=column + 1, value=headers[column])

        # Selected Status List
        # Select All if user did not made any choice
        if len(statuses) == 0:
            for key in self.cacheSignageStatus().keys():
                statuses.append(key)

        record_count = 1
        for row in range(self.sourceModel().rowCount()):
            record = self.sourceModel().record(row)
            if record.value(self.Fields.Type.index) in types:
                if record.value(self.Fields.Status.index) in statuses:
                    record_count += 1
                    values = list(map(record.value, model_fields))

                    # Signage Type
                    values[4] = self.cacheSignageType().get(record.value(self.Fields.Type.index)).name

                    # Signage Status
                    values[2] = self.cacheSignageStatus().get(record.value(self.Fields.Status.index)).name
    
                    if include_publicnote:
                        if values[-1] is not None:
                            values[-1] = html2text(values[-1]).strip()

                    ws.append(values)

        if record_count == 1:
            record_count = 2

        table = Table(displayName="Table1", ref=f"{xrange}{record_count}")
        style = TableStyleInfo(name="TableStyleMedium2",
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=False)
        table.tableStyleInfo = style

        ws.column_dimensions["B"].width = 150.0
        ws.column_dimensions["A"].number_format = numbers.FORMAT_TEXT

        # Conditional formatting
        red_text = Font(color="9C0006")
        red_fill = PatternFill(bgColor="FFC7CE")
        ws.conditional_formatting.add(f"F2:F{record_count}", CellIsRule(operator='equal', formula=['0'], stopIfTrue=False, font=red_text, fill=red_fill))

        for cell in ws['F':'F']:
            cell.alignment = Alignment(horizontal="center")

        try:
            ws.add_table(table)
            wb.save(destination)
        except Exception as e:
            err = f"Cannot save the file - The file might be opened.\nError:{e}"
            logger.error(err)
            return False, err
        
        return True, "Export successfull !"
        
    def importFromExcel(self, selected_files: list, update_title: bool = False):
        df: DataFrame = mergeExcelFiles(selected_files, drop_duplicate="first", outfile="")

        # Validate table
        try:
            df["Type"]
            df["Title"]
            df["Owner"]
            df['Refkey']
        except KeyError as err:
            msg = f"Cannot import data from Excel - KeyError: could not find {err}"
            logger.error(msg)
            return False, msg
        except Exception as err:
            msg = f"Cannot import data from Excel - Error: could not find {err}"
            logger.error(msg)
            return False, msg

        # df.fillna(' ', method='ffill', inplace=True)
        df.sort_values(by=['Refkey'], inplace=True)
        self.sort(self.Fields.Refkey.index, Qt.SortOrder.AscendingOrder)

        for _, df_row in df.iterrows():
            df_refkey = f'{df_row["Refkey"]:>{'0'}{3}}' # format rekey with leading '0'
            df_type = df_row["Type"].lower()
            df_title = df_row["Title"]
            df_owner = df_row["Owner"]

            for row in range(self.sourceModel().rowCount()):
                index = self.sourceModel().index(row, 0)
                refkey_index = self.sourceModel().index(row, self.Fields.Refkey.index)
                m_refkey = self.sourceModel().data(refkey_index, QtCore.Qt.ItemDataRole.EditRole)
                type_index = self.sourceModel().index(row, self.Fields.Type.index)
                m_signage_type = self.sourceModel().data(type_index, QtCore.Qt.ItemDataRole.EditRole)

                if self.cacheSignageType().get(df_type):
                    if m_refkey == df_refkey and m_signage_type == self.cacheSignageType().get(df_type).uid:
                        if update_title:
                            res = self.sourceModel().setData(index.sibling(row, self.Fields.Title.index), df_title, Qt.ItemDataRole.EditRole)
                            if res:
                                self.sourceModel().submit()
                        break
            else:  # Insert if not found
                if df_refkey != "":
                    try:
                        src = f'{{"application":"InspectorMate", "module":"ImportFromExcel", "item":"{*selected_files,}"}}'
                        signage = Signage(refkey=df_refkey,
                                          title=df_title,
                                          owner=df_owner,
                                          type=self.cacheSignageType().get(df_type).uid,
                                          source=src,
                                          workspace_id=self.activeWorkspace().id)
                        self.insertSignage(signage=signage, update_treemodel=False)
                    except Exception as e:
                        err = f"Error importing Signage: refkey={df_refkey} -- Error: {e}\nVerify the field 'Type'. Cannot import unknown signage type"
                        logger.error(err)
                        return False, err     
        
        ok, err = self.refreshSourceModel()
        if not ok:
            return False, err

        ok, err = self.refreshTreeModel()
        if not ok:
            return False, err
        
        return True, "Import successfull !"
    
    def summary(self) -> list:
        """Get a summary of signage's status"""
        
        #          | Request | Question | 
        # Open     |   1     |   1      |  
        # Close    |   0     |   1      |  
        # Total    |   2     |   2      |
                
        # Vertical headers
        vheaders = list(self.cacheSignageStatus().strkeys())
        vheaders.append("Total")

        # Horizontal headers
        hheaders = list(self.cacheSignageType().strkeys())

        # Init data table
        data = [[0] * len(hheaders) for i in range(len(vheaders))]  

        # Populate data table
        for row in range(self._source_model.rowCount()):
            t = self._source_model.index(row, self.Fields.Type.index).data(QtCore.Qt.ItemDataRole.DisplayRole)
            s = self._source_model.index(row, self.Fields.Status.index).data(QtCore.Qt.ItemDataRole.DisplayRole)

            if t is not None and s is not None:
                data[int(s)][int(t)] += 1    # row data
                data[-1][int(t)] += 1        # Total row

        return data, vheaders, hheaders
    
    def cacheOESignage(self):
        """Cache OneNote signage"""
        self.cache_oe_signage.clear()
        if AppDatabase.activeWorkspace().OESectionID() is None:
            return

        import json

        for row in range(self.sourceModel().rowCount()):
            source_json = self.sourceModel().index(row, self.Fields.Source.index).data(QtCore.Qt.ItemDataRole.DisplayRole)
            source: dict = json.loads(source_json)
            application = source.get("application")
            if application == "OneNote":
                object_id = source.get("object_id")
                self.cache_oe_signage.append(object_id)
        
    def loadFromOnenote(self):
        """Get a list of tag from OneNote module"""
        oe_section_id = AppDatabase.activeWorkspace().OESectionID()
        oe_section_name = AppDatabase.activeWorkspace().OESectionName()

        if oe_section_id is None:
            return True, "OneNote Section not defined"
                
        tags = OnenoteModel.fetch_onenote(oe_section_id)

        if len(tags) == 0:
            return True, "Nothing new to import"

        regex = mconf.default_regex if mconf.settings.value("regex") is None or mconf.settings.value("regex") == "" else mconf.settings.value("regex")

        tag: OE.Tag
        n = 0
        for tag in tags:
            if tag.object_id in self.cache_oe_signage:
                continue

            signage = Signage()
            signage.title = html2text(tag.text).strip()
            signage.refkey = find_match(signage.title, regex)
            signage_type: SignageType = AppDatabase.cache_signage_type.get(tag.type.lower().strip())

            # Ignore unknown signage
            if signage_type is None:
                continue
            signage_type_id = signage_type.uid

            signage.type = signage_type_id
            signage.workspace_id = AppDatabase.activeWorkspace().id
            src = f'{{"application":"OneNote", "module":"loadFromOnenote", "section":"{oe_section_name}", "page":"{tag.page_name}", "object_id":"{tag.object_id}"}}'
            signage.source = src
            signage.creation_datetime = datetime.fromisoformat(tag.creationTime[:-1]).astimezone(timezone.utc).strftime('%Y-%m-%d')
            signage.modification_datetime = tag.lastModifiedTime

            ok, err = self.insertSignage(signage=signage, update_treemodel=False)
            if not ok:
                return False, err
            
            n += 1
            self.cache_oe_signage.append(tag.object_id)

        ok, err = self.refreshTreeModel()
        if not ok:
            return False, err
            
        return True, f"Import from OneNote successfull!\n{n} tags imported"


class SignageProxyModel(ProxyModel):
    def __init__(self, model):
        super().__init__(model)

        self.setSourceModel(model)
        self.setRecursiveFilteringEnabled(True)

        self.owner_filter = []
        self.evidence_filter = False
        self.evidence_column = 0

    def setOwnerFilter(self, owners: list, column: int):
        self.owner_filter = owners
        self.owner_column = column

    def setEvidenceFilter(self, evidence_only: QtCore.Qt.CheckState, column: int):
        self.evidence_filter = evidence_only
        self.evidence_column = column

    def filterAcceptsRow(self, source_row: int, source_parent: QtCore.QModelIndex) -> bool:
        owner_filter = True
        evidence_filter = True

        # Owner filter
        if len(self.owner_filter) > 0:
            index_owner = self.sourceModel().index(source_row, self.owner_column, source_parent)
            data_owner = self.sourceModel().data(index_owner)
            if data_owner in self.owner_filter:
                owner_filter = True
            else:
                owner_filter = False

        # Evidence filter
        index_evidence = self.sourceModel().index(source_row, self.evidence_column, source_parent)
        data_evidence = self.sourceModel().data(index_evidence)
        # with evidence only
        if self.evidence_filter == QtCore.Qt.CheckState.Checked:
            evidence_filter = bool(data_evidence)
        # without evidence only
        elif self.evidence_filter == QtCore.Qt.CheckState.Unchecked:
            evidence_filter = not bool(data_evidence)
            
        return ProxyModel.filterAcceptsRow(self, source_row, source_parent) and owner_filter and evidence_filter